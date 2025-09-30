#!/usr/bin/env python3
import requests
import pandas as pd
from datetime import datetime
from dateutil import parser
from geopy.distance import geodesic

# Coordonnées centre Montpellier (Place de la Comédie)
MTP_COORDS = (43.6119, 3.8777)

def fetch_events():
    url = "https://data.montpellier3m.fr/api/records/1.0/search/"
    params = {
        "dataset": "agenda-evenements-3m@data-herault-occitanie",
        "rows": 1000,
        "facet": ["categorie", "lieu"]
    }
    resp = requests.get(url, params=params)
    resp.raise_for_status()
    data = resp.json()
    return data["records"]

def normalize(record):
    fields = record.get("fields", {})
    titre = fields.get("titre", "Sans titre")
    date_debut = fields.get("date_debut")
    date_fin = fields.get("date_fin", date_debut)
    lieu = fields.get("lieu", "Lieu inconnu")
    description = fields.get("description", "")
    gratuit = "gratuit" in description.lower() or "entrée libre" in description.lower()
    lat, lon = None, None
    if "geo_point_2d" in fields:
        lat, lon = fields["geo_point_2d"]
    return {
        "Titre": titre,
        "Date début": date_debut,
        "Date fin": date_fin,
        "Lieu": lieu,
        "Description": description,
        "Gratuit": gratuit,
        "Latitude": lat,
        "Longitude": lon,
    }

def is_inside(lat, lon):
    if not lat or not lon:
        return False
    dist = geodesic(MTP_COORDS, (lat, lon)).km
    return dist <= 5

def is_outside(lat, lon):
    if not lat or not lon:
        return False
    dist = geodesic(MTP_COORDS, (lat, lon)).km
    return 5 < dist <= 20

def export(events, filename):
    df = pd.DataFrame(events)
    colors = {
        "concert": "FF9999",
        "théâtre": "99FF99",
        "enfant": "FF99CC",
        "cinéma": "FFFF99",
        "autre": "D2B48C",
    }
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Événements"
    ws.append(df.columns.tolist())
    for _, row in df.iterrows():
        values = row.tolist()
        ws.append(values)
        cat = "autre"
        desc = (row["Description"] or "").lower()
        if "concert" in desc:
            cat = "concert"
        elif "théâtre" in desc or "theatre" in desc:
            cat = "théâtre"
        elif "enfant" in desc or "jeunesse" in desc:
            cat = "enfant"
        elif "cinéma" in desc or "film" in desc:
            cat = "cinéma"
        fill = PatternFill(start_color=colors[cat], end_color=colors[cat], fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill
    wb.save(filename)

def main():
    records = fetch_events()
    events = [normalize(r) for r in records if r]
    gratuits = [e for e in events if e["Gratuit"]]

    intra = [e for e in gratuits if is_inside(e["Latitude"], e["Longitude"])]
    outside = [e for e in gratuits if is_outside(e["Latitude"], e["Longitude"])]

    pd.DataFrame(intra).to_csv("INTRA_montpellier_gratuits.csv", index=False)
    pd.DataFrame(outside).to_csv("OUTSIDE_20km_gratuits.csv", index=False)
    export(intra + outside, "evenements_montpellier_auto.xlsx")

if __name__ == "__main__":
    main()
